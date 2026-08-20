import os
import io
import functools
import time
import cloudinary
from translations import TRANSLATIONS
import cloudinary.uploader
from flask import (
    Flask, render_template, request, jsonify,
    redirect, url_for, session, flash, send_file, abort
)
from dotenv import load_dotenv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import db as DB

load_dotenv()

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'fallback-secret-key')

# ── Cloudinary configuration ──────────────────────────────────
cloudinary.config(
    cloud_name = os.environ.get('CLOUDINARY_CLOUD_NAME'),
    api_key    = os.environ.get('CLOUDINARY_API_KEY'),
    api_secret = os.environ.get('CLOUDINARY_API_SECRET'),
    secure     = True
)

# ── Initialise DB on startup ──────────────────────────────────
# In production, table creation & seeding run only when ENABLE_DB_INIT=true
if os.environ.get('ENABLE_DB_INIT', 'false').lower() == 'true' or os.environ.get('FLASK_ENV') == 'development':
    with app.app_context():
        try:
            DB.init_db()
            DB.seed_competitions()
            DB.seed_cms_data()
        except Exception as e:
            print(f"[DB INIT WARNING] {e}")


@app.route('/api/prayer-times')
def get_prayer_times():
    import http.client
    import urllib.parse
    import json
    from flask import jsonify
    from datetime import datetime

    lat = request.args.get('latitude', '51.5194682')
    lng = request.args.get('longitude', '-0.1360365')
    tz = request.args.get('timezone', 'UTC')

    now = datetime.now()
    today_str = now.strftime('%d-%m-%Y')
    current_month = now.month
    current_year = now.year
    current_day = now.day

    try:
        # 1. Fetch Prayer Timings
        conn = http.client.HTTPSConnection("api.aladhan.com")
        path = (
            f"/v1/timings/{today_str}"
            f"?latitude={lat}&longitude={lng}"
            f"&method=3&shafaq=general"
            f"&tune=5,3,5,7,9,-1,0,8,-6"
            f"&school=0&midnightMode=0"
            f"&timezonestring={urllib.parse.quote(tz)}"
            f"&latitudeAdjustmentMethod=1&calendarMethod=UAQ&iso8601=false"
        )

        headers = {
            "Accept-Encoding": ""
        }
        conn.request("GET", path, headers=headers)
        response = conn.getresponse()
        data = response.read().decode('utf-8')
        conn.close()

        timings_data = json.loads(data)

        # 2. Fetch Islamic Calendar using UAQ calendarMethod
        conn_cal = http.client.HTTPSConnection("api.aladhan.com")
        cal_path = f"/v1/gToHCalendar/{current_month}/{current_year}?calendarMethod=UAQ"
        conn_cal.request("GET", cal_path, headers=headers)
        res_cal = conn_cal.getresponse()
        cal_data_raw = res_cal.read().decode('utf-8')
        conn_cal.close()

        calendar_data = json.loads(cal_data_raw)

        # Merge UAQ Hijri date into timings_data
        if calendar_data.get('code') == 200 and len(calendar_data.get('data', [])) >= current_day:
            today_hijri = calendar_data['data'][current_day - 1].get('hijri')
            if today_hijri and 'data' in timings_data and 'date' in timings_data['data']:
                timings_data['data']['date']['hijri'] = today_hijri

        return jsonify(timings_data)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ═══════════════════════════════════════════════════════════════
#  Helpers
# ═══════════════════════════════════════════════════════════════
ADMIN_USERNAME = os.environ.get('ADMIN_USERNAME', 'admin')
ADMIN_PASSWORD = os.environ.get('ADMIN_PASSWORD', 'AlQuranAdmin2026')

STATUS_LABELS = {
    'upcoming':  'Upcoming',
    'ongoing':   'Ongoing',
    'completed': 'Completed',
}

THEME_MAP = {
    'teal':  'comp-top--teal',
    'gold':  'comp-top--gold',
    'green': 'comp-top--green',
    'grey':  'comp-top--grey',
}

BADGE_MAP = {
    'upcoming':  'comp-badge--upcoming',
    'ongoing':   'comp-badge--ongoing',
    'completed': 'comp-badge--completed',
}


def admin_required(f):
    @functools.wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('admin_logged_in'):
            return redirect(url_for('admin_login'))
        return f(*args, **kwargs)
    return decorated


@app.context_processor
def inject_pending_reviews():
    """Inject pending review count & new enquiries count into all templates (for admin sidebar badges)."""
    try:
        if session.get('admin_logged_in'):
            cms_data = DB.get_all_homepage_cms_data()
            all_reviews = cms_data.get('all_reviews', [])
            all_enquiries = cms_data.get('enquiries', [])
            rev_cnt = sum(1 for r in all_reviews if r.get('status') == 'pending')
            enq_cnt = sum(1 for e in all_enquiries if e.get('status') == 'new')
            return {'pending_reviews_count': rev_cnt, 'new_enquiries_count': enq_cnt}
    except Exception:
        pass
    return {'pending_reviews_count': 0, 'new_enquiries_count': 0}




@app.context_processor
def inject_translations():
    """Inject current-language translation dict (`t`) and `lang` into every template."""
    lang = session.get('lang', 'en')
    t_obj = TRANSLATIONS.get(lang, TRANSLATIONS['en'])
    # Wrap in a simple attribute-access object so templates can use t.key
    class _T(dict):
        def __getattr__(self, item):
            try:
                return self[item]
            except KeyError:
                return ''
    return {'t': _T(t_obj), 'lang': lang}


@app.context_processor
def inject_active_courses():
    try:
        courses = DB.get_all_courses(visible_only=True)
    except Exception:
        courses = []
    return {'active_courses': courses}


@app.after_request
def add_performance_and_cache_headers(response):
    """Instruct browsers & CDN to cache static assets (CSS, JS, images, fonts) for 1 year."""
    if request.path.startswith('/static/'):
        response.headers['Cache-Control'] = 'public, max-age=31536000, immutable'
    return response


@app.route('/set-lang/<lang>')
def set_lang(lang):
    """Set the UI language via session cookie, then redirect back."""
    if lang in ('en', 'ur'):
        session['lang'] = lang
    return redirect(request.referrer or url_for('index'))


def _parse_tags(raw):
    """Split comma-separated tags string into cleaned list."""
    if not raw:
        return []
    return [t.strip() for t in raw.split(',') if t.strip()]


def _build_excel(registrations, title='Registrations'):
    """Build an openpyxl workbook from a list of registration dicts."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:31]  # Excel sheet name max 31 chars

    # Styles
    hdr_font    = Font(bold=True, color='FFFFFF', size=11)
    hdr_fill    = PatternFill('solid', fgColor='0A3D33')
    hdr_align   = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border_side = Side(style='thin', color='CCCCCC')
    cell_border = Border(left=border_side, right=border_side,
                         top=border_side, bottom=border_side)

    headers = ['#', 'Competition', 'Name', 'Email', 'Phone',
               'Age', 'Country', 'Experience', 'Notes', 'Registered At']
    col_widths = [5, 30, 22, 28, 18, 10, 18, 35, 35, 22]

    for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = hdr_align
        cell.border    = cell_border
        ws.column_dimensions[cell.column_letter].width = w

    ws.row_dimensions[1].height = 28

    alt_fill = PatternFill('solid', fgColor='F5F5F5')
    for row_idx, reg in enumerate(registrations, start=2):
        row_data = [
            row_idx - 1,
            reg.get('competition_title', ''),
            reg.get('name', ''),
            reg.get('email', ''),
            reg.get('phone', ''),
            reg.get('age', ''),
            reg.get('country', ''),
            reg.get('experience', ''),
            reg.get('notes', ''),
            str(reg.get('created_at', ''))[:19],
        ]
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = cell_border
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            if row_idx % 2 == 0:
                cell.fill = alt_fill
        ws.row_dimensions[row_idx].height = 20

    return wb


# ═══════════════════════════════════════════════════════════════
#  Public Routes — Courses
# ═══════════════════════════════════════════════════════════════
@app.route('/')
def index():
    try:
        comps = DB.get_all_competitions()
    except Exception:
        comps = []
    featured = [c for c in comps if c['status'] != 'completed'][:3]

    cms_data = DB.get_all_homepage_cms_data()
    settings = cms_data.get('settings', {})

    hero_setting = settings.get('hero', {})
    hero = hero_setting.get('data') or {}
    hero_visible = hero_setting.get('is_visible', True)

    about_setting = settings.get('about', {})
    about = about_setting.get('data') or {}
    about_visible = about_setting.get('is_visible', True)

    seo_setting = settings.get('seo', {})
    seo = seo_setting.get('data') or {}

    courses_hdr_setting = settings.get('courses_header', {})
    courses_hdr = courses_hdr_setting.get('data') or {}
    courses_hdr_visible = courses_hdr_setting.get('is_visible', True)

    academic_hdr_setting = settings.get('academic_header', {})
    academic_hdr = academic_hdr_setting.get('data') or {}
    academic_hdr_visible = academic_hdr_setting.get('is_visible', True)

    why_hdr_setting = settings.get('why_choose_header', {})
    why_hdr = why_hdr_setting.get('data') or {}
    why_hdr_visible = why_hdr_setting.get('is_visible', True)

    faq_hdr_setting = settings.get('faq_header', {})
    faq_hdr = faq_hdr_setting.get('data') or {}
    faq_hdr_visible = faq_hdr_setting.get('is_visible', True)

    reviews_hdr_setting = settings.get('reviews_header', {})
    reviews_hdr = reviews_hdr_setting.get('data') or {}
    reviews_hdr_visible = reviews_hdr_setting.get('is_visible', True)

    all_reviews = cms_data.get('all_reviews', [])
    reviews = [r for r in all_reviews if r.get('status') == 'approved' and r.get('is_visible', True)]

    visible_courses = [c for c in cms_data.get('courses', []) if c.get('is_visible', True)]
    visible_academic = [a for a in cms_data.get('academic_subjects', []) if a.get('is_visible', True)]
    visible_stats = [s for s in cms_data.get('stats', []) if s.get('is_visible', True)]
    visible_features = [f for f in cms_data.get('features', []) if f.get('is_visible', True)]
    visible_faqs = [q for q in cms_data.get('faqs', []) if q.get('is_visible', True)]

    return render_template('pages/index.html',
                           featured_comps=featured,
                           reviews=reviews,
                           hero=hero, hero_visible=hero_visible,
                           about=about, about_visible=about_visible,
                           seo=seo,
                           courses_hdr=courses_hdr, courses_hdr_visible=courses_hdr_visible,
                           academic_hdr=academic_hdr, academic_hdr_visible=academic_hdr_visible,
                           why_hdr=why_hdr, why_hdr_visible=why_hdr_visible,
                           faq_hdr=faq_hdr, faq_hdr_visible=faq_hdr_visible,
                           reviews_hdr=reviews_hdr, reviews_hdr_visible=reviews_hdr_visible,
                           courses=visible_courses,
                           academic_subjects=visible_academic,
                           stats=visible_stats,
                           features=visible_features,
                           faqs=visible_faqs,
                           THEME_MAP=THEME_MAP, BADGE_MAP=BADGE_MAP,
                           STATUS_LABELS=STATUS_LABELS)




def _get_course_by_keyword(key):
    try:
        courses = DB.get_all_courses(visible_only=False)
        for c in courses:
            slug = c.get('slug') or ''
            link = c.get('link_url') or ''
            title = (c.get('title') or '').lower()
            if key in slug or key in link or key in title:
                return c
    except Exception:
        pass
    return None


@app.route('/courses')
def all_courses():
    try:
        courses = DB.get_all_courses(visible_only=True)
    except Exception:
        courses = []
    
    page = request.args.get('page', 1, type=int)
    per_page = 6
    total = len(courses)
    total_pages = (total + per_page - 1) // per_page
    
    start = (page - 1) * per_page
    end = start + per_page
    paginated_courses = courses[start:end]
    
    return render_template('pages/all_courses.html', 
                           courses=paginated_courses, 
                           page=page, 
                           total_pages=total_pages)


@app.route('/course/<slug>')
def course_detail(slug):
    key = slug
    if key == 'quran-recitation':
        key = 'recitation'
        
    course = _get_course_by_keyword(key)
    if not course:
        abort(404)
        
    try:
        all_courses = DB.get_all_courses(visible_only=True)
    except Exception:
        all_courses = []
        
    other_courses_list = [c for c in all_courses if c['slug'] != course['slug']][:6]
    
    return render_template('pages/course_detail_dynamic.html', course=course, other_courses_list=other_courses_list)


@app.route('/team')
def team():
    cms_data = DB.get_all_homepage_cms_data()
    all_members = cms_data.get('team_members', [])
    team_members = [m for m in all_members if m.get('is_visible', True)]
    return render_template('pages/team.html', team_members=team_members)



# ═══════════════════════════════════════════════════════════════
#  Public Routes — Competitions
# ═══════════════════════════════════════════════════════════════
@app.route('/competitions')
def competitions():
    try:
        comps = DB.get_all_competitions()
    except Exception:
        comps = []
    return render_template('pages/competitions.html', competitions=comps,
                           THEME_MAP=THEME_MAP, BADGE_MAP=BADGE_MAP,
                           STATUS_LABELS=STATUS_LABELS)


@app.route('/competitions/<int:cid>/register', methods=['GET'])
def register_page(cid):
    comp = DB.get_competition(cid)
    if not comp:
        abort(404)
    if comp['status'] == 'completed':
        flash('This competition has ended. Registration is closed.', 'warning')
        return redirect(url_for('competitions'))
    return render_template('pages/register.html', competition=comp)


@app.route('/competitions/<int:cid>/register', methods=['POST'])
def register_submit(cid):
    comp = DB.get_competition(cid)
    if not comp or comp['status'] == 'completed':
        abort(404)

    name  = request.form.get('name', '').strip()
    email = request.form.get('email', '').strip()
    phone = request.form.get('phone', '').strip()

    if not name or not phone:
        flash('Name and phone number are required.', 'error')
        return redirect(url_for('competitions') + '#all-competitions')

    DB.create_registration({
        'competition_id':    cid,
        'competition_title': comp['title'],
        'name':       name,
        'email':      email,
        'phone':      request.form.get('phone', '').strip(),
        'age':        request.form.get('age', '').strip(),
        'country':    request.form.get('country', '').strip(),
        'experience': request.form.get('experience', '').strip(),
        'notes':      request.form.get('notes', '').strip(),
    })

    flash(f'JazakAllah Khair! Your registration for "{comp["title"]}" has been received. We\'ll be in touch soon, insha\'Allah.', 'success')
    return redirect(url_for('competitions') + '#all-competitions')


@app.route('/competitions/<int:cid>/register/success')
def register_success(cid):
    comp = DB.get_competition(cid)
    if not comp:
        abort(404)
    return render_template('pages/register.html', competition=comp, success=True)


# ═══════════════════════════════════════════════════════════════
#  Contact Form (existing)
# ═══════════════════════════════════════════════════════════════
_RECENT_ENQUIRY_SUBMISSIONS = {}

@app.route('/contact', methods=['POST'])
def contact():
    if request.is_json:
        data = request.get_json() or {}
    else:
        data = request.form.to_dict()

    name = data.get('name', '').strip()
    phone = data.get('phone', '').strip()

    if name and phone:
        # Deduplication safeguard (ignore duplicate requests within 10 seconds)
        sub_key = f"{phone.lower()}:{name.lower()}"
        now = time.time()
        last_time = _RECENT_ENQUIRY_SUBMISSIONS.get(sub_key, 0)
        if now - last_time < 10:
            return jsonify({'status': 'ok',
                            'message': "JazakAllah Khair! We will contact you within 24 hours, insha'Allah."})
        _RECENT_ENQUIRY_SUBMISSIONS[sub_key] = now

        try:
            DB.create_contact_enquiry({
                'name': name,
                'phone': phone,
                'email': data.get('email', '').strip(),
                'course': data.get('course', '').strip(),
                'age': str(data.get('age', '')).strip(),
                'address': data.get('address', '').strip(),
                'message': data.get('message', '').strip()
            })
        except Exception as e:
            print("Error storing contact enquiry:", e)

    return jsonify({'status': 'ok',
                    'message': "JazakAllah Khair! We will contact you within 24 hours, insha'Allah."})




# ═══════════════════════════════════════════════════════════════
#  Public — Review Submission
# ═══════════════════════════════════════════════════════════════
@app.route('/reviews/submit', methods=['POST'])
def review_submit():
    name  = request.form.get('name', '').strip()
    text  = request.form.get('review_text', '').strip()
    if not name or not text:
        flash('Name and review text are required.', 'error')
        return redirect(url_for('index') + '#reviews')
    try:
        rating = max(1, min(5, int(request.form.get('rating', 5))))
    except (ValueError, TypeError):
        rating = 5
    DB.create_review({
        'name':        name,
        'location':    request.form.get('location', '').strip(),
        'course':      request.form.get('course', '').strip(),
        'rating':      rating,
        'review_text': text,
    })
    flash("JazakAllah Khair! Your review has been submitted and will appear after admin approval, insha'Allah.", 'success')
    return redirect(url_for('index') + '#reviews')


# ═══════════════════════════════════════════════════════════════
#  Admin — Authentication
# ═══════════════════════════════════════════════════════════════
@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    if session.get('admin_logged_in'):
        return redirect(url_for('admin_competitions'))

    error = None
    if request.method == 'POST':
        if (request.form.get('username') == ADMIN_USERNAME and
                request.form.get('password') == ADMIN_PASSWORD):
            session['admin_logged_in'] = True
            return redirect(url_for('admin_competitions'))
        error = 'Invalid username or password.'

    return render_template('admin/login.html', error=error)


@app.route('/admin/logout')
def admin_logout():
    session.pop('admin_logged_in', None)
    return redirect(url_for('admin_login'))


@app.route('/admin')
@app.route('/admin/')
@admin_required
def admin_dashboard():
    return redirect(url_for('admin_homepage'))


# ═══════════════════════════════════════════════════════════════
#  Admin — Homepage CMS & Subject Management
# ═══════════════════════════════════════════════════════════════
@app.route('/admin/homepage')
@admin_required
def admin_homepage():
    tab = request.args.get('tab', 'hero')
    cms_data = DB.get_all_homepage_cms_data()
    settings = cms_data.get('settings', {})

    return render_template('admin/homepage.html',
                           tab=tab,
                           hero_setting=settings.get('hero', {}),
                           about_setting=settings.get('about', {}),
                           seo_setting=settings.get('seo', {}),
                           courses_hdr_setting=settings.get('courses_header', {}),
                           academic_hdr_setting=settings.get('academic_header', {}),
                           why_hdr_setting=settings.get('why_choose_header', {}),
                           faq_hdr_setting=settings.get('faq_header', {}),
                           reviews_hdr_setting=settings.get('reviews_header', {}),
                           courses=cms_data.get('courses', []),
                           academic_subjects=cms_data.get('academic_subjects', []),
                           stats=cms_data.get('stats', []),
                           features=cms_data.get('features', []),
                           faqs=cms_data.get('faqs', []),
                           all_reviews=cms_data.get('all_reviews', []))



@app.route('/admin/subjects')
@app.route('/admin/courses')
@admin_required
def admin_subjects_shortcut():
    return redirect(url_for('admin_homepage', tab='courses'))


@app.route('/admin/homepage/settings/<key>', methods=['POST'])
@admin_required
def admin_homepage_save_settings(key):
    valid_keys = {'hero', 'about', 'seo', 'courses_header', 'academic_header', 'why_choose_header', 'faq_header', 'reviews_header'}
    if key not in valid_keys:
        flash('Invalid setting section', 'error')
        return redirect(url_for('admin_homepage'))

    existing = DB.get_homepage_setting(key)
    current_data = existing.get('data') if isinstance(existing.get('data'), dict) else {}

    is_visible = request.form.get('is_visible') in ('on', 'true', '1')

    if key == 'hero':
        image_url = current_data.get('image_url')
        if 'image' in request.files and request.files['image'].filename:
            try:
                res = cloudinary.uploader.upload(request.files['image'], folder='alquran/cms')
                image_url = res.get('secure_url')
            except Exception as e:
                flash(f'Hero image upload failed: {e}', 'warning')
        new_data = {
            'arabic_heading': request.form.get('arabic_heading', '').strip(),
            'title': request.form.get('title', '').strip(),
            'sub_title': request.form.get('sub_title', '').strip(),
            'badge': request.form.get('badge', '').strip(),
            'description': request.form.get('description', '').strip(),
            'primary_cta_text': request.form.get('primary_cta_text', '').strip(),
            'primary_cta_url': request.form.get('primary_cta_url', '').strip(),
            'secondary_cta_text': request.form.get('secondary_cta_text', '').strip(),
            'secondary_cta_url': request.form.get('secondary_cta_url', '').strip(),
            'image_url': image_url or request.form.get('image_url', '')
        }
    elif key == 'about':
        image_url = current_data.get('image_url')
        if 'image' in request.files and request.files['image'].filename:
            try:
                res = cloudinary.uploader.upload(request.files['image'], folder='alquran/cms')
                image_url = res.get('secure_url')
            except Exception as e:
                flash(f'About image upload failed: {e}', 'warning')
        raw_points = request.form.get('mission_points', '')
        points = [p.strip() for p in raw_points.split('\n') if p.strip()]
        new_data = {
            'tag': request.form.get('tag', '').strip(),
            'title': request.form.get('title', '').strip(),
            'title_gold': request.form.get('title_gold', '').strip(),
            'description': request.form.get('description', '').strip(),
            'mission_points': points,
            'image_url': image_url or request.form.get('image_url', '')
        }
    elif key == 'seo':
        og_image = current_data.get('og_image')
        if 'og_image_file' in request.files and request.files['og_image_file'].filename:
            try:
                res = cloudinary.uploader.upload(request.files['og_image_file'], folder='alquran/cms')
                og_image = res.get('secure_url')
            except Exception as e:
                flash(f'OG image upload failed: {e}', 'warning')
        new_data = {
            'title': request.form.get('title', '').strip(),
            'meta_description': request.form.get('meta_description', '').strip(),
            'og_image': og_image or request.form.get('og_image', ''),
            'canonical_url': request.form.get('canonical_url', '').strip()
        }
    else: # Header titles
        new_data = {
            'tag': request.form.get('tag', '').strip(),
            'title': request.form.get('title', '').strip(),
            'title_span': request.form.get('title_span', '').strip(),
            'description': request.form.get('description', '').strip()
        }

    DB.save_homepage_setting(key, new_data, is_visible=is_visible)
    flash(f'{key.replace("_", " ").title()} section updated successfully!', 'success')
    return redirect(url_for('admin_homepage', tab=key if key in ('hero','about','seo') else 'courses'))


# ── Courses Management ──
@app.route('/admin/homepage/courses/save', methods=['POST'])
@admin_required
def admin_courses_save():
    cid = request.form.get('id', type=int)
    image_url = request.form.get('image_url', '')
    if 'image' in request.files and request.files['image'].filename:
        try:
            res = cloudinary.uploader.upload(request.files['image'], folder='alquran/courses')
            image_url = res.get('secure_url')
        except Exception as e:
            flash(f'Course image upload failed: {e}', 'warning')

    raw_badges = request.form.get('badges', '')
    badges = [b.strip() for b in raw_badges.split(',') if b.strip()]

    title = request.form['title'].strip()
    slug = request.form.get('slug', '').strip()
    if not slug:
        import re
        slug = re.sub(r'[^a-z0-9]+', '-', title.lower()).strip('-')

    link_url = request.form.get('link_url', '').strip()
    if not link_url:
        link_url = f"/course/{slug}"

    data = {
        'title': title,
        'slug': slug,
        'category': request.form.get('category', 'Quranic Studies').strip(),
        'description': request.form.get('description', '').strip(),
        'image_url': image_url,
        'badges': badges,
        'link_url': link_url,
        'seo_title': request.form.get('seo_title', '').strip(),
        'seo_description': request.form.get('seo_description', '').strip(),
        'display_order': request.form.get('display_order', 0, type=int),
        'is_visible': request.form.get('is_visible') in ('on', 'true', '1')
    }

    if cid:
        DB.update_course(cid, data)
        flash('Course updated successfully!', 'success')
    else:
        DB.create_course(data)
        flash('New course created successfully!', 'success')

    return redirect(url_for('admin_homepage', tab='courses'))


# ── Team Management ──
@app.route('/admin/team')
@admin_required
def admin_team():
    cms_data = DB.get_all_homepage_cms_data()
    team_members = cms_data.get('team_members', [])
    return render_template('admin/team.html', team_members=team_members)


@app.route('/admin/team/save', methods=['POST'])
@admin_required
def admin_team_save():
    tid = request.form.get('id', type=int)
    image_url = request.form.get('image_url', '')
    if 'image' in request.files and request.files['image'].filename:
        try:
            res = cloudinary.uploader.upload(request.files['image'], folder='alquran/team')
            image_url = res.get('secure_url')
        except Exception as e:
            flash(f'Team image upload failed: {e}', 'warning')

    data = {
        'name': request.form['name'].strip(),
        'role_label': request.form.get('role_label', '').strip(),
        'subject': request.form.get('subject', '').strip(),
        'education': request.form.get('education', '').strip(),
        'image_url': image_url or request.form.get('image_url', ''),
        'is_director': request.form.get('is_director') in ('on', 'true', '1'),
        'display_order': request.form.get('display_order', 0, type=int),
        'is_visible': request.form.get('is_visible') in ('on', 'true', '1')
    }

    if tid:
        DB.update_team_member(tid, data)
        flash('Team member updated successfully!', 'success')
    else:
        DB.create_team_member(data)
        flash('New team member added successfully!', 'success')

    return redirect(url_for('admin_team'))


@app.route('/admin/team/<int:tid>/delete', methods=['POST'])
@admin_required
def admin_team_delete(tid):
    DB.delete_team_member(tid)
    flash('Team member deleted successfully!', 'success')
    return redirect(url_for('admin_team'))



@app.route('/admin/homepage/courses/<int:cid>/delete', methods=['POST'])
@admin_required
def admin_courses_delete(cid):
    DB.delete_course(cid)
    flash('Course deleted.', 'info')
    return redirect(url_for('admin_homepage', tab='courses'))


@app.route('/admin/courses/<int:cid>/details', methods=['GET', 'POST'])
@admin_required
def admin_course_details(cid):
    course = DB.get_course(cid)
    if not course:
        abort(404)
        
    details = course.get('course_details') or {}
    if isinstance(details, str):
        import json
        try:
            details = json.loads(details)
        except Exception:
            details = {}
    course['course_details'] = details

    if request.method == 'POST':
        why_items = [x.strip() for x in request.form.get('why_items', '').split('\n') if x.strip()]
        learn_items = [x.strip() for x in request.form.get('learn_items', '').split('\n') if x.strip()]
        cert_items = [x.strip() for x in request.form.get('cert_items', '').split('\n') if x.strip()]

        new_details = {
            "course_arabic": request.form.get('course_arabic', '').strip(),
            "course_tagline": request.form.get('course_tagline', '').strip(),
            "course_duration": request.form.get('course_duration', '').strip(),
            "course_schedule": request.form.get('course_schedule', '').strip(),
            "course_eligibility": request.form.get('course_eligibility', '').strip(),
            "course_certificate": request.form.get('course_certificate', '').strip(),
            "course_mode": request.form.get('course_mode', 'Online (Live)').strip(),
            "course_fee": request.form.get('course_fee', 'Contact Us').strip(),
            "quote_text": request.form.get('quote_text', '').strip(),
            "quote_cite": request.form.get('quote_cite', '').strip(),
            "intro_text": request.form.get('intro_text', '').strip(),
            "why_items_header": request.form.get('why_items_header', '').strip(),
            "why_items": why_items,
            "learn_header": request.form.get('learn_header', 'What You Will Learn').strip(),
            "learn_items": learn_items,
            "highlights_header": request.form.get('highlights_header', 'Course Highlights').strip(),
            "highlights": [
                {
                    "icon": request.form.get('hl1_icon', 'hl_instructor.svg').strip(),
                    "title": request.form.get('hl1_title', '').strip(),
                    "desc": request.form.get('hl1_desc', '').strip()
                },
                {
                    "icon": request.form.get('hl2_icon', 'hl_feedback.svg').strip(),
                    "title": request.form.get('hl2_title', '').strip(),
                    "desc": request.form.get('hl2_desc', '').strip()
                },
                {
                    "icon": request.form.get('hl3_icon', 'hl_global.svg').strip(),
                    "title": request.form.get('hl3_title', '').strip(),
                    "desc": request.form.get('hl3_desc', '').strip()
                },
                {
                    "icon": request.form.get('hl4_icon', 'hl_certificate.svg').strip(),
                    "title": request.form.get('hl4_title', '').strip(),
                    "desc": request.form.get('hl4_desc', '').strip()
                }
            ],
            "why_header": request.form.get('why_header', '').strip(),
            "why_p1": request.form.get('why_p1', '').strip(),
            "why_p2": request.form.get('why_p2', '').strip(),
            "cert_header": request.form.get('cert_header', 'Certification Requirements').strip(),
            "cert_items": cert_items
        }

        course['course_details'] = new_details
        DB.update_course(cid, course)
        flash(f'Course details for "{course["title"]}" updated successfully!', 'success')
        return redirect(url_for('admin_homepage', tab='courses'))

    return render_template('admin/course_details.html', course=course)


# ── Academic Subjects Management ──
@app.route('/admin/homepage/academic/save', methods=['POST'])
@admin_required
def admin_academic_save():
    aid = request.form.get('id', type=int)
    icon_or_image = request.form.get('icon_or_image', '📚')
    if 'image' in request.files and request.files['image'].filename:
        try:
            res = cloudinary.uploader.upload(request.files['image'], folder='alquran/academic')
            icon_or_image = res.get('secure_url')
        except Exception as e:
            flash(f'Academic subject image upload failed: {e}', 'warning')

    raw_badges = request.form.get('badges', '')
    badges = [b.strip() for b in raw_badges.split(',') if b.strip()]

    data = {
        'title': request.form['title'].strip(),
        'description': request.form.get('description', '').strip(),
        'icon_or_image': icon_or_image,
        'badges': badges,
        'cta_text': request.form.get('cta_text', 'Enroll via Contact').strip(),
        'cta_url': request.form.get('cta_url', '#contact').strip(),
        'display_order': request.form.get('display_order', 0, type=int),
        'is_visible': request.form.get('is_visible') in ('on', 'true', '1')
    }

    if aid:
        DB.update_academic_subject(aid, data)
        flash('Academic subject updated successfully!', 'success')
    else:
        DB.create_academic_subject(data)
        flash('Academic subject created successfully!', 'success')

    return redirect(url_for('admin_homepage', tab='academic'))


@app.route('/admin/homepage/academic/<int:aid>/delete', methods=['POST'])
@admin_required
def admin_academic_delete(aid):
    DB.delete_academic_subject(aid)
    flash('Academic subject deleted.', 'info')
    return redirect(url_for('admin_homepage', tab='academic'))


# ── Statistics Management ──
@app.route('/admin/homepage/stats/save', methods=['POST'])
@admin_required
def admin_stats_save():
    sid = request.form.get('id', type=int)
    data = {
        'value': request.form['value'].strip(),
        'label': request.form['label'].strip(),
        'icon': request.form.get('icon', '📊').strip(),
        'display_order': request.form.get('display_order', 0, type=int),
        'is_visible': request.form.get('is_visible') in ('on', 'true', '1')
    }
    if sid:
        DB.update_statistic(sid, data)
        flash('Statistic updated.', 'success')
    else:
        DB.create_statistic(data)
        flash('Statistic created.', 'success')

    return redirect(url_for('admin_homepage', tab='stats'))


@app.route('/admin/homepage/stats/<int:sid>/delete', methods=['POST'])
@admin_required
def admin_stats_delete(sid):
    DB.delete_statistic(sid)
    flash('Statistic deleted.', 'info')
    return redirect(url_for('admin_homepage', tab='stats'))


# ── Features Management (Why Choose Us) ──
@app.route('/admin/homepage/features/save', methods=['POST'])
@admin_required
def admin_features_save():
    fid = request.form.get('id', type=int)
    image_url = request.form.get('image_url', '')
    if 'image' in request.files and request.files['image'].filename:
        try:
            res = cloudinary.uploader.upload(request.files['image'], folder='alquran/features')
            image_url = res.get('secure_url')
        except Exception as e:
            flash(f'Feature image upload failed: {e}', 'warning')

    data = {
        'title': request.form['title'].strip(),
        'description': request.form.get('description', '').strip(),
        'image_url': image_url,
        'icon': request.form.get('icon', '⭐').strip(),
        'display_order': request.form.get('display_order', 0, type=int),
        'is_visible': request.form.get('is_visible') in ('on', 'true', '1')
    }
    if fid:
        DB.update_feature(fid, data)
        flash('Feature updated.', 'success')
    else:
        DB.create_feature(data)
        flash('Feature created.', 'success')

    return redirect(url_for('admin_homepage', tab='features'))


@app.route('/admin/homepage/features/<int:fid>/delete', methods=['POST'])
@admin_required
def admin_features_delete(fid):
    DB.delete_feature(fid)
    flash('Feature deleted.', 'info')
    return redirect(url_for('admin_homepage', tab='features'))


# ── FAQs Management ──
@app.route('/admin/homepage/faqs/save', methods=['POST'])
@admin_required
def admin_faqs_save():
    fid = request.form.get('id', type=int)
    data = {
        'question': request.form['question'].strip(),
        'answer': request.form['answer'].strip(),
        'display_order': request.form.get('display_order', 0, type=int),
        'is_visible': request.form.get('is_visible') in ('on', 'true', '1')
    }
    if fid:
        DB.update_faq(fid, data)
        flash('FAQ updated.', 'success')
    else:
        DB.create_faq(data)
        flash('FAQ created.', 'success')

    return redirect(url_for('admin_homepage', tab='faqs'))


@app.route('/admin/homepage/faqs/<int:fid>/delete', methods=['POST'])
@admin_required
def admin_faqs_delete(fid):
    DB.delete_faq(fid)
    flash('FAQ deleted.', 'info')
    return redirect(url_for('admin_homepage', tab='faqs'))



# ═══════════════════════════════════════════════════════════════
#  Admin — Competitions
# ═══════════════════════════════════════════════════════════════
@app.route('/admin/competitions')
@admin_required
def admin_competitions():
    comps = DB.get_all_competitions()
    # attach registration counts
    for c in comps:
        c['reg_count'] = DB.get_registration_count(c['id'])
    total_regs = sum(c['reg_count'] for c in comps)
    return render_template('admin/competitions_list.html',
                           competitions=comps,
                           total_regs=total_regs,
                           STATUS_LABELS=STATUS_LABELS)


@app.route('/admin/competitions/new', methods=['GET', 'POST'])
@admin_required
def admin_competition_new():
    if request.method == 'POST':
        image_url = None
        if 'image' in request.files and request.files['image'].filename:
            try:
                result = cloudinary.uploader.upload(
                    request.files['image'],
                    folder='alquran/competitions',
                    transformation=[{'width': 800, 'crop': 'limit'}]
                )
                image_url = result.get('secure_url')
            except Exception as e:
                flash(f'Image upload failed: {e}', 'warning')

        DB.create_competition({
            'title':        request.form['title'],
            'category':     request.form.get('category', 'Quranic Studies'),
            'description':  request.form.get('description'),
            'date_display': request.form.get('date_display'),
            'status':       request.form.get('status', 'upcoming'),
            'location':     request.form.get('location', 'Online — Worldwide'),
            'age_group':    request.form.get('age_group', 'All Ages'),
            'prize':        request.form.get('prize'),
            'tags':         _parse_tags(request.form.get('tags', '')),
            'icon':         request.form.get('icon', '🏆'),
            'color_theme':  request.form.get('color_theme', 'teal'),
            'image_url':    image_url or request.form.get('image_url'),
        })
        flash('Competition created successfully!', 'success')
        return redirect(url_for('admin_competitions'))

    return render_template('admin/competition_form.html',
                           action='new', competition=None,
                           STATUS_LABELS=STATUS_LABELS)


@app.route('/admin/competitions/<int:cid>/edit', methods=['GET', 'POST'])
@admin_required
def admin_competition_edit(cid):
    comp = DB.get_competition(cid)
    if not comp:
        abort(404)

    if request.method == 'POST':
        image_url = comp.get('image_url')
        if 'image' in request.files and request.files['image'].filename:
            try:
                result = cloudinary.uploader.upload(
                    request.files['image'],
                    folder='alquran/competitions',
                    transformation=[{'width': 800, 'crop': 'limit'}]
                )
                image_url = result.get('secure_url')
            except Exception as e:
                flash(f'Image upload failed: {e}', 'warning')

        DB.update_competition(cid, {
            'title':        request.form['title'],
            'category':     request.form.get('category', 'Quranic Studies'),
            'description':  request.form.get('description'),
            'date_display': request.form.get('date_display'),
            'status':       request.form.get('status', 'upcoming'),
            'location':     request.form.get('location', 'Online — Worldwide'),
            'age_group':    request.form.get('age_group', 'All Ages'),
            'prize':        request.form.get('prize'),
            'tags':         _parse_tags(request.form.get('tags', '')),
            'icon':         request.form.get('icon', '🏆'),
            'color_theme':  request.form.get('color_theme', 'teal'),
            'image_url':    image_url or request.form.get('image_url'),
        })
        flash('Competition updated successfully!', 'success')
        return redirect(url_for('admin_competitions'))

    # Prepare tags as comma string for the form
    comp['tags_str'] = ', '.join(comp.get('tags') or [])
    return render_template('admin/competition_form.html',
                           action='edit', competition=comp,
                           STATUS_LABELS=STATUS_LABELS)


@app.route('/admin/competitions/<int:cid>/status', methods=['POST'])
@admin_required
def admin_competition_status(cid):
    new_status = request.form.get('status')
    if new_status not in STATUS_LABELS:
        return jsonify({'error': 'Invalid status'}), 400
    DB.update_competition_status(cid, new_status)
    return jsonify({'ok': True, 'status': new_status,
                    'label': STATUS_LABELS[new_status]})


@app.route('/admin/competitions/<int:cid>/delete', methods=['POST'])
@admin_required
def admin_competition_delete(cid):
    DB.delete_competition(cid)
    flash('Competition deleted.', 'info')
    return redirect(url_for('admin_competitions'))


# ═══════════════════════════════════════════════════════════════
#  Admin — Registrations
# ═══════════════════════════════════════════════════════════════
@app.route('/admin/registrations')
@admin_required
def admin_registrations():
    cid   = request.args.get('competition_id', type=int)
    regs  = DB.get_all_registrations(competition_id=cid)
    comps = DB.get_all_competitions()
    selected_comp = DB.get_competition(cid) if cid else None
    return render_template('admin/registrations.html',
                           registrations=regs,
                           competitions=comps,
                           selected_comp=selected_comp,
                           selected_id=cid)


@app.route('/admin/registrations/<int:rid>/delete', methods=['POST'])
@admin_required
def admin_registration_delete(rid):
    cid = request.form.get('competition_id', type=int)
    DB.delete_registration(rid)
    flash('Registration deleted.', 'info')
    if cid:
        return redirect(url_for('admin_registrations', competition_id=cid))
    return redirect(url_for('admin_registrations'))


@app.route('/admin/registrations/export')
@admin_required
def admin_registrations_export():
    cid   = request.args.get('competition_id', type=int)
    regs  = DB.get_all_registrations(competition_id=cid)

    if cid:
        comp = DB.get_competition(cid)
        sheet_title = (comp['title'][:28] if comp else 'Competition')
        filename    = f"registrations_{cid}.xlsx"
    else:
        sheet_title = 'All Registrations'
        filename    = 'all_registrations.xlsx'

    wb = _build_excel(regs, title=sheet_title)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


# ═══════════════════════════════════════════════════════════════
#  Admin — Reviews Moderation
# ═══════════════════════════════════════════════════════════════
@app.route('/admin/reviews')
@admin_required
def admin_reviews():
    reviews = DB.get_all_reviews()
    pending = [r for r in reviews if r['status'] == 'pending']
    return render_template('admin/reviews.html',
                           reviews=reviews,
                           pending_count=len(pending),
                           pending_reviews_count=len(pending))


@app.route('/admin/reviews/<int:rid>/status', methods=['POST'])
@admin_required
def admin_review_status(rid):
    status = request.form.get('status')
    if status not in ('approved', 'rejected', 'pending'):
        return jsonify({'error': 'invalid status'}), 400
    DB.update_review_status(rid, status)
    return jsonify({'ok': True, 'status': status})


@app.route('/admin/reviews/<int:rid>/delete', methods=['POST'])
@admin_required
def admin_review_delete(rid):
    DB.delete_review(rid)
    flash('Review deleted.', 'success')
    return redirect(url_for('admin_reviews'))


# ═══════════════════════════════════════════════════════════════
#  Admin — Contact Enquiries Management
# ═══════════════════════════════════════════════════════════════
def _build_enquiries_excel(enquiries):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Enquiries"
    headers = ["#", "Date & Time", "Name", "Phone", "Email", "Requested Course", "Age", "Address", "Message", "Status"]
    ws.append(headers)

    header_fill = openpyxl.styles.PatternFill(start_color="0A3D33", end_color="0A3D33", fill_type="solid")
    header_font = openpyxl.styles.Font(color="FFFFFF", bold=True)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")

    for i, e in enumerate(enquiries, 1):
        dt_str = str(e.get('created_at', ''))[:19]
        ws.append([
            i,
            dt_str,
            e.get('name', ''),
            e.get('phone', ''),
            e.get('email', ''),
            e.get('course', ''),
            e.get('age', ''),
            e.get('address', ''),
            e.get('message', ''),
            e.get('status', 'new').title()
        ])

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    return wb


@app.route('/admin/enquiries')
@admin_required
def admin_enquiries():
    all_enquiries = DB.get_all_contact_enquiries()
    status_filter = request.args.get('status', 'all')
    if status_filter != 'all':
        enquiries = [e for e in all_enquiries if e.get('status') == status_filter]
    else:
        enquiries = all_enquiries

    new_count = sum(1 for e in all_enquiries if e.get('status') == 'new')
    contacted_count = sum(1 for e in all_enquiries if e.get('status') == 'contacted')

    return render_template('admin/enquiries.html',
                           enquiries=enquiries,
                           total_count=len(all_enquiries),
                           new_count=new_count,
                           contacted_count=contacted_count,
                           selected_status=status_filter)


@app.route('/admin/enquiries/<int:eid>/status', methods=['POST'])
@admin_required
def admin_enquiry_status(eid):
    status = request.form.get('status', 'contacted')
    if status not in ('new', 'contacted', 'archived'):
        return jsonify({'error': 'invalid status'}), 400
    DB.update_enquiry_status(eid, status)
    flash(f'Enquiry #{eid} status updated to {status}.', 'success')
    return redirect(request.referrer or url_for('admin_enquiries'))


@app.route('/admin/enquiries/<int:eid>/delete', methods=['POST'])
@admin_required
def admin_enquiry_delete(eid):
    DB.delete_enquiry(eid)
    flash('Enquiry deleted successfully.', 'info')
    return redirect(url_for('admin_enquiries'))


@app.route('/admin/enquiries/export')
@admin_required
def admin_enquiries_export():
    enquiries = DB.get_all_contact_enquiries()
    wb = _build_enquiries_excel(enquiries)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='contact_enquiries.xlsx'
    )


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)


